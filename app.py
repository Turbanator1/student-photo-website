from flask import Flask, render_template, request, send_file
import pandas as pd
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font
import zipfile
import os
import tempfile
import traceback # This lets us catch and print errors!

app = Flask(__name__)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/merge', methods=['POST'])
def merge():
    try: # START OF SAFETY NET
        # 1. Get the uploaded files
        csv_file = request.files.get('csv_file')
        zip_file = request.files.get('zip_file')

        if not csv_file or not zip_file:
            return "Error: Please upload both files.", 400

        # 2. Create a temporary directory to handle the processing safely
        temp_dir = tempfile.mkdtemp()
        
        # 3. Save and extract the zip file
        zip_path = os.path.join(temp_dir, "photos.zip")
        zip_file.save(zip_path)
        
        extract_folder = os.path.join(temp_dir, "extracted_photos")
        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
            zip_ref.extractall(extract_folder)

        # 4. Read the CSV (with fallback encoding for Excel-generated CSVs)
        try:
            df = pd.read_csv(csv_file, encoding='utf-8')
        except UnicodeDecodeError:
            csv_file.seek(0)
            df = pd.read_csv(csv_file, encoding='latin1')
        
        # 5. Create new Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Merged Data"

        # Set up centering and bold fonts
        center_align = Alignment(horizontal='center', vertical='center')
        bold_font = Font(bold=True)

        # 6. Clean up the headers
        clean_headers = []
        for col in df.columns:
            col_name = str(col).strip()
            if col_name.lower() in ["fathername", "father name"]: col_name = "Father Name"
            elif col_name.lower() in ["fathermobile", "father mobile"]: col_name = "Father Mobile"
            elif col_name.lower() in ["rollno", "roll no", "roll"]: col_name = "Roll No"
            elif col_name.lower() == "name": col_name = "Name"
            elif col_name.lower() == "class": col_name = "Class"
            elif col_name.lower() == "section": col_name = "Section"
            clean_headers.append(col_name)

        headers = clean_headers + ["Photo"]
        ws.append(headers)

        for cell in ws[1]:
            cell.alignment = center_align
            cell.font = bold_font

        # 7. Process each student
        for index, row in df.iterrows():
            # Skip rows that are completely empty
            if row.isnull().all():
                continue

            ws.append(list(row))
            current_row = ws.max_row 
            
            for cell in ws[current_row]:
                cell.alignment = center_align
            
            raw_roll = str(row.iloc[0]).strip() 
            
            # Skip if the roll number cell is blank (NaN)
            if raw_roll.lower() == 'nan' or not raw_roll:
                continue
            
            try:
                search_roll = str(int(float(raw_roll)))
            except ValueError:
                search_roll = raw_roll 
            
            # Look for the image in the extracted folder
            img_path = None
            for root, dirs, files in os.walk(extract_folder):
                for file in files:
                    file_name_only = os.path.splitext(file)[0].strip()
                    try:
                        clean_file_name = str(int(float(file_name_only)))
                    except ValueError:
                        clean_file_name = file_name_only
                    
                    if clean_file_name == search_roll and file.lower().endswith(('.png', '.jpg', '.jpeg')):
                        img_path = os.path.join(root, file)
                        break
                if img_path:
                    break
            
            # Insert image into Excel
            if img_path:
                try:
                    img = Image(img_path)
                    img.width, img.height = 80, 80
                    cell_letter = openpyxl.utils.get_column_letter(len(headers))
                    cell_name = f"{cell_letter}{current_row}"
                    ws.add_image(img, cell_name)
                    ws.row_dimensions[current_row].height = 65
                except Exception as img_err:
                    print(f"Warning: Could not add image for Roll No {search_roll}. Reason: {img_err}")

        # Adjust the width of the final Photo column
        ws.column_dimensions[openpyxl.utils.get_column_letter(len(headers))].width = 15

        # 8. Save the final file
        output_path = os.path.join(temp_dir, "Final_Student_Data.xlsx")
        wb.save(output_path)

        # Send the file to the user to download
        return send_file(output_path, as_attachment=True, download_name="Final_Student_Data.xlsx")

    # END OF SAFETY NET
    except Exception as e:
        # If anything crashes, print it to terminal and send error to browser!
        error_details = traceback.format_exc()
        print("CRASH LOG:\n", error_details)
        return f"""
        <div style='font-family: sans-serif; text-align: center; margin-top: 50px;'>
            <h1 style='color: red;'>An Error Occurred!</h1>
            <p>The system failed while trying to process the files. Reason:</p>
            <p><b>{str(e)}</b></p>
            <p><i>Check your command prompt/terminal for the full crash report.</i></p>
            <button onclick='window.history.back()' style='padding: 10px 20px; font-size: 16px; cursor: pointer;'>Go Back</button>
        </div>
        """, 500

if __name__ == '__main__':
    app.run(debug=True)